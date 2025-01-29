import os
import django
import openpyxl
import re
from django.db import transaction

# Configura el entorno de Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'tu_proyecto.settings')
django.setup()

from .models import Curso, TNivel, RequisitoBase, RequisitoCreditos,RequisitoCursos, Formula, PeriodoAcademico, Horario
from calificacion.models import Competencia,SubCompetencia
from usuarios.models import Profesor,Usuario, TRol
from django.db.models import ObjectDoesNotExist
from django.db.utils import IntegrityError
from faker import Faker
from datetime import date
import random

fake = Faker()


# Función para extraer el número de horas desde la celda
def extraer_numero(celda):
    if isinstance(celda, str):
        return int(''.join(filter(str.isdigit, celda)))
    return celda  # En caso de que no sea un string

def mapear_nivel(nivel):
    nivel_str = str(nivel).strip()  # Convertir el nivel a string y eliminar espacios

    # Verificar si el nivel es un número entre 1 y 10
    if nivel_str.isdigit() and nivel_str in dict(TNivel.choices):
        return nivel_str  # Retornar el nivel como string
    print(f"NivelStr: {nivel_str} y Nivel: {nivel}")
    
    # Verificar si es 'E' o cualquier otra opción válida
    if int(nivel_str) == 0:
        return TNivel.ELECTIVO

    # Si no coincide con ningún valor válido, lanzar un error
    raise ValueError(f"Nivel '{nivel}' no es válido.")

# Función para procesar el campo 'requisitos' de cada curso
def procesar_requisitos(requisito, curso):
    requisito = requisito.strip().lower()

    # Caso 1: Si es un requisito de créditos (ejemplo: "44 créditos")
    if re.match(r"^\d+\s*créditos?$", requisito):
        total_creditos = float(re.search(r"\d+", requisito).group())
        requisito_creditos, _ = RequisitoCreditos.objects.get_or_create(total_creditos=total_creditos)
        curso.requisitos = requisito_creditos

    # Caso 2: Si no tiene requisitos ("no tiene")
    elif requisito == "no tiene":
        requisito_creditos, _ = RequisitoCreditos.objects.get_or_create(total_creditos=0)
        curso.requisitos = requisito_creditos

    # Caso 3: Si tiene una lista de cursos como requisito (ejemplo: "1MAT09, 1MAT10")
    else:
        # Separar las claves y buscar los cursos correspondientes
        claves_cursos = [clave.strip() for clave in requisito.split(',')]
        cursos_requisito = Curso.objects.filter(clave__in=claves_cursos)

        if cursos_requisito.exists():
            # **Crear un nuevo RequisitoCursos cada vez que se detecta una combinación de cursos**
            requisito_cursos = RequisitoCursos.objects.create()
            requisito_cursos.cursos.add(*cursos_requisito)  # Agregar los cursos
            requisito_cursos.save()  # Guardar la instancia en la base de datos

            # Asignar el requisito creado al curso actual
            curso.requisitos = requisito_cursos
        else:
            print(f"Advertencia: No se encontraron cursos para las claves: {claves_cursos}")

    # Guardar el curso con el requisito asignado
    curso.save()


# Función principal para cargar los cursos desde Excel
def cargar_cursos_desde_excel():
    EXCEL_PATH = r"E:\DOCUMENTOS\Documentos PUCP\12avo Ciclo\IngeSoft\c.1inf37.24.2.Plan.de.Estudios.xlsx"
    
    # Cargar el archivo Excel y seleccionar la hoja activa
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active

    #rangos = [(17, 19)]  # Ajusta según tus necesidades
    rangos = [(4,35),(38,85),(88,126)]
    
    # Iterar sobre cada rango definido
    with transaction.atomic():
        for inicio, fin in rangos:
            print(f"Procesando filas desde {inicio} hasta {fin}...")

            # Iterar sobre el rango especificado
            for row in sheet.iter_rows(min_row=inicio, max_row=fin, min_col=2, max_col=18, values_only=True):
                # Extraer datos principales del curso
                clave, nombre, nivel, creditos, tipo, sesiones, ME, *competencias, requisitos = row

                # Validación: si 'clave' es None o está vacía, se salta la fila
                if not clave or str(clave).strip() == "":
                    print("Fila ignorada: Clave vacía")
                    continue  # Saltar a la siguiente fila
                
                if clave == "Clave":
                    continue
                
                # Crear la instancia del curso
                curso = Curso.objects.create(
                    clave=clave,
                    nombre=nombre,
                    nivel=mapear_nivel(int(nivel)),
                    creditos=float(creditos),
                    obligatorio=tipo.strip().startswith("01"),
                    numHoras=extraer_numero(sesiones),
                    formula=Formula.objects.get(id=ME),
                    activo=True
                )

                # Asignar las competencias relacionadas
                for idx, competencia_value in enumerate(competencias, start=1):
                    print(f"IDX: {idx}  CompetenciaValue: {competencia_value}")
                    if competencia_value == 'x':  # Si hay una 'X', se agrega la competencia
                        clave_competencia = f"C{idx}"  # Generar la clave 'Cn' donde n es el índice
                        print(f"Competencia: {clave_competencia}")
                        # Crear o recuperar la competencia
                        competencia, created = Competencia.objects.get_or_create(
                            clave=clave_competencia,
                            defaults={'nombre': f'Competencia {idx}', 'descripcion': ''}
                        )
                        # Añadir la competencia al curso
                        curso.competencias.add(competencia)

                # Guardar el curso para asegurar que la relación ManyToMany se aplique
                curso.save()

                # Procesar y asignar los requisitos
                procesar_requisitos(requisitos, curso)
                
                print(f"Curso '{curso.nombre}' cargado exitosamente.")

            print("Todos los cursos han sido cargados.")
        
        
        
#################################################3
##### Cargar HORARIOS
        
def generar_telefono_formateado():
    # Generar las tres partes del número
    parte1 = random.randint(900, 999)  # Ej: 987
    parte2 = random.randint(100, 999)  # Ej: 654
    parte3 = random.randint(100, 999)  # Ej: 321

    # Formatear como '987 654 321'
    telefono = f"{parte1} {parte2} {parte3}"
    return telefono

#Crea los profesores de los horarios
def crear_profesor_si_no_existe(clave_profesor):
    try:
        profesor = Profesor.objects.get(codigo=clave_profesor)
    except ObjectDoesNotExist:
        # Crear un nuevo usuario usando Faker
        nombre = fake.first_name()
        primer_apellido = fake.last_name()
        segundo_apellido = fake.last_name()
        base_correo = f"{nombre.lower()}.{primer_apellido.lower()}@example.com"
        codigo = f"{clave_profesor}"  # Código único del profesor

        # Intentar crear el profesor y manejar duplicados de correo
        profesor = None
        intentos = 0
        while profesor is None and intentos < 5:
            correo = base_correo if intentos == 0 else f"{nombre.lower()}.{primer_apellido.lower()}{random.randint(1, 100)}@example.com"
            try:
                profesor = Profesor.objects.create(
                    nombre=nombre,
                    primerApellido=primer_apellido,
                    segundoApellido=segundo_apellido,
                    correo=correo,
                    codigo=codigo,
                    fechaRegistro=date.today(),
                    telefono=generar_telefono_formateado(),
                    activo=True,
                    tipo=TRol.REGULAR
                )
                print(f"Profesor {clave_profesor} creado con el usuario {profesor.nombre} y correo {correo}")
            except IntegrityError:
                print(f"El correo {correo} ya existe. Intentando con una variante...")
                intentos += 1

        if profesor is None:
            raise Exception(f"No se pudo crear el profesor {clave_profesor} después de varios intentos.")

    return profesor

# Función principal para cargar los horarios desde Excel
def cargar_horarios_desde_excel():
    EXCEL_PATH = r"E:\DOCUMENTOS\Documentos PUCP\12avo Ciclo\IngeSoft\c.1inf37.24.2.Horarios.clases.xlsx"
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active  # Seleccionar la hoja activa

    horarios_previos = 0  # Guardar el número de horarios anterior si falta en la fila
    vacantes_previas = 0  # Guardar las vacantes anteriores si falta en la fila

    # Recorrer las filas del Excel
    for row in sheet.iter_rows(min_row=4, max_row=120, values_only=True):
        clave_curso = row[1]  # Columna B - Clave del curso

        if not clave_curso:
            # Si la clave del curso está vacía, saltar esta fila
            continue

        # Buscar el curso por su clave
        try:
            curso = Curso.objects.get(clave=clave_curso)
        except ObjectDoesNotExist:
            print(f"El curso con clave {clave_curso} no existe.")
            continue

        # Obtener el número de horarios (columna E) y vacantes (columna F)
        horarios = row[4] if row[4] is not None else horarios_previos
        vacantes = row[5] if row[5] is not None else vacantes_previas

        # Actualizar las variables previas en caso de celdas vacías
        horarios_previos = horarios
        vacantes_previas = vacantes

        # Procesar horarios y profesores
        for col_horario, col_profesor in zip(range(7, 27), range(28, 48)):  # H:AA y AC:AV
            clave_horario = row[col_horario]
            clave_profesor = row[col_profesor]
            
            if not clave_horario or not clave_profesor:
                continue  # Si no hay clave de horario o profesor, pasar a la siguiente

            if clave_horario == '---' or clave_horario==r"'---":
                continue

            # Crear o buscar el horario
            horario, _ = Horario.objects.get_or_create(
                claveHorario=clave_horario,
                idCurso=curso,
                defaults={
                    'numVacantes': vacantes,
                    'numMatriculados': 0,
                    'numAprobados': 0,
                    'numDesaprobados': 0
                }
            )

            # Crear el profesor si no existe
            profesor = crear_profesor_si_no_existe(clave_profesor)

            # Asignar el profesor al horario
            horario.idprofesor = profesor  # Asegúrate de asignar la instancia completa, no solo el código
            horario.save()  # Guardar los cambios en la base de datos

            print(f"Horario {clave_horario} asignado a profesor {clave_profesor} para curso {curso.clave}")

    print("Carga de horarios completada.")

        
        
        
# Ejecución del script desde consola
if __name__ == "__main__":
    cargar_cursos_desde_excel()